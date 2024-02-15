import os
from dotenv import load_dotenv
import requests
from urllib3.exceptions import InsecureRequestWarning
from openpyxl import load_workbook
import openpyxl
import concurrent.futures
import smtplib
import xml.etree.ElementTree as ET


#Loading environment variables
load_dotenv()
PASSWORD = os.environ.get("PASSWORD")
# FILENAME = os.environ.get("FILENAME")
user_input = input("Please enter xls filename from XML Files directory to be used: ")
FILENAME = f'./XML_FILES/{user_input}'

# disable ssl warning
requests.packages.urllib3.disable_warnings(InsecureRequestWarning)

#Importing XML File
file_path = f"./XML_Files/{input('Enter file name: ')}"
tree = ET.parse(file_path)
root = tree.getroot()
command_string = ET.tostring(root, encoding='unicode', method='xml', xml_declaration=None, default_namespace=None, short_empty_elements=True)
print(command_string)

#Defining command function
DEFAULT_PASSWORD = 'Basic ' + PASSWORD

def command_request(auth_val=DEFAULT_PASSWORD, ip=None, idx=None, file=None):
    url = f"https://{ip}/putxml"
    cell_no = idx+2
    payload = command_string

    headers = {'Content-Type': 'text/xml','Authorization': auth_val}
    try:
        response = requests.request("POST", url, headers=headers, data=payload, verify=False)
        response.raise_for_status()
        print(response.status_code)
        file[f"D{cell_no}"] = response.status_code
    except requests.exceptions.HTTPError as err:
        print(err.response.status_code)
        file[f"D{cell_no}"] = err.response.status_code

def send_commands(excel_file):
    try:
        #Loading excel workbook with IPs
        codecList = load_workbook(f"./codec_lists/{excel_file}")
        codecSheet = codecList.active
        with concurrent.futures.ThreadPoolExecutor(max_workers=5) as executor:
            for idx, value in enumerate(codecSheet.iter_rows(min_row=2, min_col=3, max_col=3, values_only=True)):
                ip = value[0]
                executor.submit(command_request, DEFAULT_PASSWORD, ip, idx, codecList)
    except Exception as error:
        print(error)

send_commands(FILENAME)